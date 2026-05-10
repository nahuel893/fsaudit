"""Excel reporter — generates multi-sheet .xlsx workbooks via openpyxl."""

from __future__ import annotations

import csv as csv_mod
import math
from collections import defaultdict
from pathlib import Path

from datetime import datetime
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from fsaudit.analyzer.metrics import AnalysisResult
from fsaudit.reporter.base import BaseReporter
from fsaudit.scanner.models import FileRecord

if TYPE_CHECKING:
    from fsaudit.security.models import SecurityResult

# Excel row limits.
EXCEL_MAX_ROWS: int = 1_048_576
MAX_INVENTORY_ROWS: int = EXCEL_MAX_ROWS - 1  # 1,048,575 data rows

# Sheet names in required order (7 fixed sheets; inventory is dynamic).
SHEET_NAMES: list[str] = [
    "Dashboard",
    "Por Categoria",
    "Timeline",
    "Top Archivos Pesados",
    "Archivos Inactivos",
    "Alertas",
    "Por Directorio",
]

# Severity order for sorting (lower index = higher severity)
_SEVERITY_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3}

# Severity fill colors for the Seguridad sheet
_SEVERITY_FILLS = {
    "critical": PatternFill(fill_type="solid", fgColor="FFFF0000"),  # red
    "high": PatternFill(fill_type="solid", fgColor="FFFFA500"),      # orange
    "medium": PatternFill(fill_type="solid", fgColor="FFFFFF00"),    # yellow
    "low": PatternFill(fill_type="solid", fgColor="FFD3D3D3"),       # light gray
}


class ExcelReporter(BaseReporter):
    """Concrete reporter that writes an 8-sheet (or 9-sheet with security) Excel workbook."""

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def __init__(self, overflow_strategy: str = "shard") -> None:
        self.overflow_strategy = overflow_strategy

    def _inventory_sheet_names(self, n_records: int) -> list[str]:
        """Return inventory sheet names based on record count.

        When ``n_records <= MAX_INVENTORY_ROWS``, returns a single
        ``"Inventario Completo"``. Otherwise returns N shard names
        like ``"Inventario 1/3"``, ``"Inventario 2/3"``, ...
        """
        if n_records <= MAX_INVENTORY_ROWS:
            return ["Inventario Completo"]
        n_shards = math.ceil(n_records / MAX_INVENTORY_ROWS)
        return [f"Inventario {i} de {n_shards}" for i in range(1, n_shards + 1)]

    def _write_inventory_sheets(
        self, wb: Workbook, records: list[FileRecord]
    ) -> int:
        """Create and populate inventory sheet(s), one per shard if needed.

        Returns the number of inventory sheets created (1 for normal case,
        N for sharded case).
        """
        names = self._inventory_sheet_names(len(records))
        n_shards = len(names)

        for i, name in enumerate(names):
            ws = wb.create_sheet(title=name)
            chunk_start = i * MAX_INVENTORY_ROWS
            chunk_end = chunk_start + MAX_INVENTORY_ROWS
            chunk = records[chunk_start:chunk_end]
            self._write_inventario_chunk(ws, chunk)

        return n_shards

    def generate(
        self,
        records: list[FileRecord],
        analysis: AnalysisResult,
        output_path: Path,
        *,
        security: "SecurityResult | None" = None,
    ) -> Path:
        """Create report at *output_path*.

        When ``overflow_strategy == "csv"``, writes a UTF-8 BOM CSV file
        instead of an xlsx workbook.  Otherwise (``"shard"``), produces
        the standard multi-sheet Excel workbook with dynamic inventory
        sharding when rows exceed ``MAX_INVENTORY_ROWS``.

        Args:
            records: Classified file records.
            analysis: Pre-computed analysis metrics.
            output_path: Destination file path. Parent dir must exist.
            security: Optional security scan result.

        Raises:
            FileNotFoundError: If the parent directory of *output_path*
                does not exist.
        """
        output_path = Path(output_path)
        if not output_path.parent.exists():
            raise FileNotFoundError(
                f"Parent directory does not exist: {output_path.parent}"
            )

        # CSV mode: generate CSV instead of xlsx, then return early.
        if self.overflow_strategy == "csv":
            return self._generate_csv(records, output_path)

        wb = Workbook()

        # Create fixed sheets in order — first sheet is auto-created by Workbook().
        for idx, name in enumerate(SHEET_NAMES):
            if idx == 0:
                wb.active.title = name  # type: ignore[union-attr]
            else:
                wb.create_sheet(title=name)

        # Create inventory sheet(s) dynamically (shard if overflow, single otherwise).
        inventory_count = self._write_inventory_sheets(wb, records)

        # Detect overflow and build warning state.
        # Overflow occurs when inventory_count > 1 (sharded) or always for CSV.
        overflow = inventory_count > 1
        self._overflow_warning: str | None = None
        if overflow:
            total = len(records)
            self._overflow_warning = (
                f"Inventario dividido en {inventory_count} hojas "
                f"({total} archivos totales). Estrategia: shard"
            )

        # Delegate writing to private methods.
        self._write_dashboard(
            wb[SHEET_NAMES[0]], analysis, records,
            security=security,
            overflow=overflow,
            inventory_count=inventory_count,
        )
        self._write_por_categoria(wb[SHEET_NAMES[1]], analysis)
        self._write_timeline(wb[SHEET_NAMES[2]], analysis)
        self._write_top_pesados(wb[SHEET_NAMES[3]], analysis)
        self._write_inactivos(wb[SHEET_NAMES[4]], analysis)
        self._write_alertas(wb[SHEET_NAMES[5]], analysis, records)
        self._write_por_directorio(wb[SHEET_NAMES[6]], records)

        # Optional 9th sheet — only when security data is present
        if security is not None:
            seg_ws = wb.create_sheet(title="Seguridad")
            self._write_security_sheet(seg_ws, security)

        wb.save(str(output_path))
        return output_path

    # ------------------------------------------------------------------
    # Styling helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _parse_period(period: str) -> datetime:
        """Convert a YYYY-MM-01 string to a datetime for date axis."""
        return datetime.strptime(period, "%Y-%m-%d")

    @staticmethod
    def _bytes_to_mb(size_bytes: int) -> float:
        """Convert bytes to MB, rounded to 2 decimals."""
        return size_bytes / (1024 * 1024)

    @staticmethod
    def _format_bytes(size_bytes: int) -> str:
        """Return human-readable byte string (B / KB / MB / GB)."""
        if size_bytes < 0:
            return f"{size_bytes} B"
        for unit, threshold in [("GB", 1 << 30), ("MB", 1 << 20), ("KB", 1 << 10)]:
            if size_bytes >= threshold:
                return f"{size_bytes / threshold:.2f} {unit}"
        return f"{size_bytes} B"

    @staticmethod
    def _apply_header_style(ws: Worksheet, num_cols: int) -> None:
        """Bold the first row and freeze panes at row 2."""
        bold = Font(bold=True)
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = bold
        ws.freeze_panes = "A2"

    @staticmethod
    def _apply_autofilter(ws: Worksheet, num_cols: int) -> None:
        """Apply autofilter from A1 to last data cell."""
        last_col = get_column_letter(num_cols)
        last_row = ws.max_row
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    @staticmethod
    def _auto_column_width(
        ws: Worksheet,
        max_width: int = 50,
        sample_rows: int = 100,
    ) -> None:
        """Adjust column widths based on content (sampled)."""
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells[:sample_rows + 1]:  # +1 for header
                try:
                    length = len(str(cell.value)) if cell.value is not None else 0
                except Exception:
                    length = 0
                if length > max_len:
                    max_len = length
            adjusted = min(max_len + 2, max_width)
            ws.column_dimensions[col_letter].width = max(adjusted, 8)

    # ------------------------------------------------------------------
    # Sheet writers
    # ------------------------------------------------------------------

    def _write_dashboard(
        self,
        ws: Worksheet,
        analysis: AnalysisResult,
        records: list[FileRecord],
        *,
        security: "SecurityResult | None" = None,
        overflow: bool = False,
        inventory_count: int = 1,
    ) -> None:
        """KPI overview sheet with professional layout.

        Layout:
          Cols A-C: KPIs + tables (left panel)
          Cols E-L: Timeline chart (right panel, immediately visible)
          Below both: Top 5 Categories + Top 5 Directories

        When *security* is provided, two additional KPI cards are appended:
        - Card 13: "Security Score" (color-coded ≥80 green, 60-79 yellow, <60 red)
        - Card 14: "Hallazgos" (count of findings)

        When *overflow* is True, two hazard rows are inserted at rows 2-3:
        - Row 2: "⚠ OVERFLOW" + warning message
        - Row 3: "Estrategia" + strategy name ("shard")
        Standard KPIs shift down by 2 rows.
        """
        title_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=12)
        value_font = Font(bold=True, size=14)
        label_font = Font(bold=True)

        # Row 1: Title
        ws.cell(row=1, column=1, value="Dashboard").font = title_font

        # Overflow warning rows (only when shard mode overflow triggered)
        if overflow:
            warning_font = Font(bold=True, size=12, color="FF0000")
            total = len(records)
            overflow_msg = (
                f"Inventario dividido en {inventory_count} hojas "
                f"({total} archivos totales)"
            )
            ws.cell(row=2, column=1, value="⚠ OVERFLOW").font = warning_font
            ws.cell(row=2, column=2, value=overflow_msg).font = warning_font
            ws.cell(row=3, column=1, value="Estrategia").font = label_font
            ws.cell(row=3, column=2, value="shard").font = value_font

        # Alert count: zero-byte + permission issues + duplicate groups
        alert_count = (
            len(analysis.zero_byte_files)
            + len(analysis.permission_issues)
            + sum(
                len(paths) for paths in analysis.duplicates_by_name.values()
            )
        )

        # Health score
        score = analysis.health_score
        score_color = "00B050" if score >= 70 else "FFC000" if score >= 40 else "FF0000"

        # KPI rows (label col A, value col B)
        kpis: list[tuple[str, object]] = [
            ("Health Score", f"{score:.1f} / 100"),
            ("Total Archivos", analysis.total_files),
            ("Tamaño Total (MB)", self._bytes_to_mb(analysis.total_size_bytes)),
            ("Categorías", len(analysis.by_category)),
            ("Alertas Activas", alert_count),
            ("Archivos Duplicados", sum(len(p) for p in analysis.duplicates_by_name.values())),
            ("Archivos Inactivos", f"{len(analysis.inactive_files)} ({round(len(analysis.inactive_files) / max(analysis.total_files, 1) * 100, 1)}%)"),
            ("Archivos 0 bytes", len(analysis.zero_byte_files)),
            ("Tamaño Promedio (MB)", self._bytes_to_mb(int(analysis.total_size_bytes / max(analysis.total_files, 1)))),
            ("Directorios Vacíos", len(analysis.empty_directories)),
        ]

        # Extension más común
        ext_counts: dict[str, int] = defaultdict(int)
        for rec in records:
            ext_counts[rec.extension or "(sin ext)"] += 1
        if ext_counts:
            top_ext = max(ext_counts.items(), key=lambda x: x[1])
            kpis.append(("Extensión Más Común", f"{top_ext[0]} ({top_ext[1]})"))

        # Archivo más pesado
        if analysis.top_largest:
            biggest = analysis.top_largest[0]
            biggest_name = Path(biggest.get("path", "")).name
            biggest_mb = self._bytes_to_mb(biggest.get("size_bytes", 0))
            kpis.append(("Archivo Más Pesado", f"{biggest_name} ({biggest_mb} MB)"))

        # Security KPIs — only when security data is present (Cards 13 & 14)
        security_kpi_rows: list[tuple[str, object, str]] = []  # (label, value, color)
        if security is not None:
            sec_score = security.security_score
            if sec_score >= 80:
                sec_color = "00B050"  # green
            elif sec_score >= 60:
                sec_color = "FFC000"  # yellow
            else:
                sec_color = "FF0000"  # red

            finding_count = len(security.findings)
            if finding_count == 0:
                hallazgos_color = "00B050"  # green
            elif finding_count <= 10:
                hallazgos_color = "FFC000"  # yellow
            else:
                hallazgos_color = "FF0000"  # red

            security_kpi_rows = [
                ("Security Score", sec_score, sec_color),
                ("Hallazgos", finding_count, hallazgos_color),
            ]

        kpi_start_row = 4 if overflow else 2
        for i, (label, value) in enumerate(kpis, start=kpi_start_row):
            ws.cell(row=i, column=1, value=label).font = label_font
            cell = ws.cell(row=i, column=2, value=value)
            cell.font = value_font
            # Color the health score value
            if label == "Health Score":
                cell.font = Font(bold=True, size=14, color=score_color)

        # Append security KPI rows after the standard KPIs
        sec_start_row = len(kpis) + (4 if overflow else 2)
        for j, (label, value, color) in enumerate(security_kpi_rows):
            row = sec_start_row + j
            ws.cell(row=row, column=1, value=label).font = label_font
            cell = ws.cell(row=row, column=2, value=value)
            cell.font = Font(bold=True, size=14, color=color)
            cell.fill = PatternFill(fill_type="solid", fgColor="FF" + color)

        total_kpi_count = len(kpis) + len(security_kpi_rows)
        current_row = total_kpi_count + (5 if overflow else 3)  # after KPIs + title + overflow + blank

        # Top 5 Categorías por Tamaño section
        ws.cell(row=current_row, column=1, value="Top 5 Categorías por Tamaño").font = section_font
        current_row += 1
        ws.cell(row=current_row, column=1, value="Categoría").font = label_font
        ws.cell(row=current_row, column=2, value="Cantidad").font = label_font
        ws.cell(row=current_row, column=3, value="Tamaño (MB)").font = label_font
        current_row += 1

        sorted_cats = sorted(
            analysis.by_category.items(),
            key=lambda x: x[1].get("bytes", 0),
            reverse=True,
        )[:5]
        for cat, stats in sorted_cats:
            ws.cell(row=current_row, column=1, value=cat)
            ws.cell(row=current_row, column=2, value=stats.get("count", 0))
            ws.cell(row=current_row, column=3, value=self._bytes_to_mb(stats.get("bytes", 0)))
            current_row += 1

        current_row += 1

        # Top 5 Directorios por Cantidad section
        ws.cell(row=current_row, column=1, value="Top 5 Directorios por Cantidad").font = section_font
        current_row += 1
        ws.cell(row=current_row, column=1, value="Directorio").font = label_font
        ws.cell(row=current_row, column=2, value="Cantidad").font = label_font
        current_row += 1

        dir_counts: dict[str, int] = defaultdict(int)
        for rec in records:
            dir_counts[rec.parent_dir] += 1

        sorted_dirs = sorted(dir_counts.items(), key=lambda x: x[1], reverse=True)[:5]
        for dirname, count in sorted_dirs:
            ws.cell(row=current_row, column=1, value=dirname)
            ws.cell(row=current_row, column=2, value=count)
            current_row += 1

        # Apply header style and column width
        ws.freeze_panes = "A2"
        self._auto_column_width(ws)

    def _write_por_categoria(
        self, ws: Worksheet, analysis: AnalysisResult
    ) -> None:
        """Category breakdown sheet."""
        headers = [
            "Categoría",
            "Cantidad",
            "Volumen (MB)",
            "% del Total",
            "Promedio (MB)",
            "Más Reciente",
            "Más Antiguo",
        ]
        ws.append(headers)
        for cat, stats in analysis.by_category.items():
            ws.append([
                cat,
                stats.get("count", 0),
                self._bytes_to_mb(stats.get("bytes", 0)),
                round(stats.get("percent", 0.0), 2),
                self._bytes_to_mb(int(stats.get("avg_size", 0.0))),
                str(stats.get("newest", "")),
                str(stats.get("oldest", "")),
            ])

        self._apply_header_style(ws, len(headers))
        self._apply_autofilter(ws, len(headers))
        self._auto_column_width(ws)

    def _write_timeline(self, ws: Worksheet, analysis: AnalysisResult) -> None:
        """Monthly distribution sheet, sorted chronologically."""
        headers = ["Período", "Cantidad"]
        ws.append(headers)
        sorted_periods = sorted(analysis.timeline.keys())
        for period in sorted_periods:
            cell = ws.cell(row=ws.max_row + 1, column=1, value=self._parse_period(period))
            cell.number_format = "yyyy-mm-dd"
            ws.cell(row=ws.max_row, column=2, value=analysis.timeline[period])

        self._apply_header_style(ws, len(headers))
        self._apply_autofilter(ws, len(headers))
        self._auto_column_width(ws)


    def _write_top_pesados(
        self, ws: Worksheet, analysis: AnalysisResult
    ) -> None:
        """Top largest files sheet."""
        headers = ["Ruta", "Nombre", "Tamaño (MB)", "Categoría", "Última Modificación"]
        ws.append(headers)
        for item in analysis.top_largest:
            path = item.get("path", "")
            ws.append([
                path,
                Path(path).name,
                self._bytes_to_mb(item.get("size_bytes", 0)),
                item.get("category", ""),
                str(item.get("mtime", "")),
            ])

        self._apply_header_style(ws, len(headers))
        self._apply_autofilter(ws, len(headers))
        self._auto_column_width(ws)

    def _write_inactivos(
        self, ws: Worksheet, analysis: AnalysisResult
    ) -> None:
        """Inactive files sheet, sorted by days_inactive descending."""
        headers = [
            "Ruta",
            "Nombre",
            "Tamaño (MB)",
            "Categoría",
            "Última Modificación",
            "Días Inactivo",
        ]
        ws.append(headers)
        sorted_items = sorted(
            analysis.inactive_files,
            key=lambda x: x.get("days_inactive", 0),
            reverse=True,
        )
        for item in sorted_items:
            path = item.get("path", "")
            ws.append([
                path,
                Path(path).name,
                self._bytes_to_mb(item.get("size_bytes", 0)),
                item.get("category", ""),
                str(item.get("mtime", "")),
                item.get("days_inactive", 0),
            ])

        self._apply_header_style(ws, len(headers))
        self._apply_autofilter(ws, len(headers))
        self._auto_column_width(ws)

    def _write_alertas(
        self,
        ws: Worksheet,
        analysis: AnalysisResult,
        records: list[FileRecord],
    ) -> None:
        """Alerts sheet — aggregates 4 alert sources."""
        headers = ["Tipo Alerta", "Nombre", "Ruta", "Detalle"]
        ws.append(headers)

        # Zero-byte files
        for item in analysis.zero_byte_files:
            path = item.get("path", "")
            ws.append(["0 bytes", Path(path).name, path, "Archivo de 0 bytes"])

        # Permission issues
        for item in analysis.permission_issues:
            path = item.get("path", "")
            issue = item.get("issue", "")
            ws.append([
                f"Permisos: {issue}",
                Path(path).name,
                path,
                f"Permisos: {item.get('permissions', '')}",
            ])

        # Duplicate filenames
        for name, paths in analysis.duplicates_by_name.items():
            for dup_path in paths:
                ws.append(["Duplicado", name, dup_path, f"{len(paths)} copias"])

        # Files with empty extension
        for rec in records:
            if rec.extension == "":
                ws.append([
                    "Sin extension",
                    rec.name,
                    str(rec.path),
                    "Archivo sin extension",
                ])

        self._apply_header_style(ws, len(headers))
        self._apply_autofilter(ws, len(headers))
        self._auto_column_width(ws)

    def _write_por_directorio(
        self, ws: Worksheet, records: list[FileRecord]
    ) -> None:
        """Top directories by volume sheet."""
        headers = [
            "Directorio",
            "Cantidad Archivos",
            "Volumen Total (MB)",
            "Volumen Promedio (MB)",
        ]
        ws.append(headers)

        # Group by parent_dir
        dir_stats: dict[str, dict] = defaultdict(
            lambda: {"count": 0, "bytes": 0}
        )
        for rec in records:
            entry = dir_stats[rec.parent_dir]
            entry["count"] += 1
            entry["bytes"] += rec.size_bytes

        # Sort by volume desc, take top 50
        sorted_dirs = sorted(
            dir_stats.items(), key=lambda x: x[1]["bytes"], reverse=True
        )[:50]

        for dirname, stats in sorted_dirs:
            avg = stats["bytes"] / stats["count"] if stats["count"] else 0
            ws.append([
                dirname,
                stats["count"],
                self._bytes_to_mb(stats["bytes"]),
                self._bytes_to_mb(int(avg)),
            ])

        self._apply_header_style(ws, len(headers))
        self._apply_autofilter(ws, len(headers))
        self._auto_column_width(ws)

    def _write_inventario(
        self, ws: Worksheet, records: list[FileRecord]
    ) -> None:
        """Complete inventory sheet with autofilter.

        Delegates to ``_write_inventario_chunk`` for the actual writing.
        """
        self._write_inventario_chunk(ws, records)

    def _write_inventario_chunk(
        self, ws: Worksheet, records: list[FileRecord]
    ) -> None:
        """Write inventory headers + data + styling to a single worksheet.

        Used both by the single-sheet case (via ``_write_inventario``)
        and by ``_write_inventory_sheets`` for each shard.
        """
        headers = [
            "Ruta", "Nombre", "Extensión", "Tamaño (MB)",
            "Categoría", "Fecha Modificación", "Fecha Creación",
            "Último Acceso", "Profundidad", "Oculto",
            "Permisos", "Directorio Padre", "Autor",
        ]
        ws.append(headers)

        for rec in records:
            ws.append([
                str(rec.path), rec.name, rec.extension,
                self._bytes_to_mb(rec.size_bytes), rec.category,
                str(rec.mtime), str(rec.creation_time),
                str(rec.atime), rec.depth, rec.is_hidden,
                rec.permissions or "", rec.parent_dir,
                rec.author or "",
            ])

        # Autofilter spanning all columns
        if records:
            last_col = get_column_letter(len(headers))
            last_row = len(records) + 1  # +1 for header
            ws.auto_filter.ref = f"A1:{last_col}{last_row}"
        else:
            last_col = get_column_letter(len(headers))
            ws.auto_filter.ref = f"A1:{last_col}1"

        self._apply_header_style(ws, len(headers))
        self._auto_column_width(ws)

    def _generate_csv(
        self, records: list[FileRecord], output_path: Path
    ) -> Path:
        """Generate a UTF-8 BOM CSV file with all inventory records.

        CSV mode replaces xlsx generation entirely — no workbook is created.
        The output filename follows ``{folder}_inventory_{date}.csv`` instead
        of the xlsx ``{folder}_audit_{date}.xlsx`` convention.

        Args:
            records: Classified file records.
            output_path: Originally intended xlsx path (stem and extension
                are replaced to produce the CSV path).

        Returns:
            Path to the created CSV file.
        """
        # Derive CSV path: change stem pattern and extension.
        # xlsx pattern: {folder}_audit_{date}.xlsx → csv pattern: {folder}_inventory_{date}.csv
        csv_name = output_path.name.replace("_audit_", "_inventory_")
        if csv_name.endswith(".xlsx"):
            csv_name = csv_name[:-5] + ".csv"
        # Handle the unlikely case where the extension isn't .xlsx
        elif not csv_name.endswith(".csv"):
            csv_name = csv_name.rsplit(".", 1)[0] + ".csv"
        csv_path = output_path.parent / csv_name

        headers = [
            "Ruta", "Nombre", "Extensión", "Tamaño (MB)",
            "Categoría", "Fecha Modificación", "Fecha Creación",
            "Último Acceso", "Profundidad", "Oculto",
            "Permisos", "Directorio Padre", "Autor",
        ]

        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv_mod.writer(f)
            writer.writerow(headers)
            for rec in records:
                writer.writerow([
                    str(rec.path), rec.name, rec.extension,
                    self._bytes_to_mb(rec.size_bytes), rec.category,
                    str(rec.mtime), str(rec.creation_time),
                    str(rec.atime), rec.depth, rec.is_hidden,
                    rec.permissions or "", rec.parent_dir,
                    rec.author or "",
                ])

        # CSV mode always triggers a warning (strategy is explicitly "csv").
        total = len(records)
        self._overflow_warning = (
            f"CSV mode: {total} archivos exportados a {csv_path.name}. "
            f"Estrategia: csv"
        )

        return csv_path

    def _write_security_sheet(
        self,
        ws: Worksheet,
        security: "SecurityResult",
    ) -> None:
        """Write the optional 9th 'Seguridad' sheet with security findings.

        Columns:
            Archivo, Regla, Severidad, Detector, Línea, Contexto

        Rows are sorted by severity DESC (critical first) then by file path.
        Severity cells are color-coded. Autofilter and freeze panes applied.
        """
        headers = ["Archivo", "Regla", "Severidad", "Detector", "Línea", "Contexto"]
        ws.append(headers)

        # Sort findings: severity DESC (critical=0, high=1, medium=2, low=3), then path
        sorted_findings = sorted(
            security.findings,
            key=lambda f: (
                _SEVERITY_ORDER.get(f.severity.value if hasattr(f.severity, "value") else str(f.severity), 99),
                str(f.path),
            ),
        )

        sev_col_idx = headers.index("Severidad") + 1  # 1-based

        for finding in sorted_findings:
            sev_str = finding.severity.value if hasattr(finding.severity, "value") else str(finding.severity)
            ws.append([
                str(finding.path),
                finding.rule_id,
                sev_str,
                finding.detector,
                finding.line_no,
                finding.match_context,
            ])
            # Color the severity cell
            fill = _SEVERITY_FILLS.get(sev_str)
            if fill:
                ws.cell(row=ws.max_row, column=sev_col_idx).fill = fill

        # Autofilter on the header row
        last_col = get_column_letter(len(headers))
        last_row = max(ws.max_row, 1)
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

        # Freeze panes at row 2
        ws.freeze_panes = "A2"

        self._apply_header_style(ws, len(headers))
        self._auto_column_width(ws)
