"""Tests for the Excel reporter module."""

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from fsaudit.analyzer.metrics import AnalysisResult
from fsaudit.reporter.base import BaseReporter
from fsaudit.reporter.excel_reporter import ExcelReporter, SHEET_NAMES
from fsaudit.scanner.models import FileRecord


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_record(
    name: str = "file.txt",
    extension: str = ".txt",
    size_bytes: int = 1024,
    category: str = "Codigo",
    parent_dir: str = "/tmp/project",
    permissions: str | None = "644",
    is_hidden: bool = False,
    depth: int = 1,
    mtime: datetime | None = None,
) -> FileRecord:
    """Helper to build a FileRecord with sensible defaults."""
    mt = mtime or datetime(2025, 6, 15, 10, 0, 0)
    return FileRecord(
        path=Path(parent_dir) / name,
        name=name,
        extension=extension,
        size_bytes=size_bytes,
        mtime=mt,
        creation_time=datetime(2025, 1, 1),
        atime=datetime(2025, 6, 20),
        depth=depth,
        is_hidden=is_hidden,
        permissions=permissions,
        category=category,
        parent_dir=parent_dir,
    )


@pytest.fixture
def sample_records() -> list[FileRecord]:
    """A small set of diverse records for testing."""
    return [
        _make_record("readme.md", ".md", 2048, "Codigo", "/tmp/a"),
        _make_record("photo.jpg", ".jpg", 5_000_000, "Multimedia", "/tmp/b"),
        _make_record("report.xlsx", ".xlsx", 30_000, "Oficina", "/tmp/a"),
        _make_record("script.py", ".py", 500, "Codigo", "/tmp/a"),
        _make_record("Makefile", "", 0, "Codigo", "/tmp/a"),  # zero-byte, no ext
    ]


@pytest.fixture
def sample_analysis() -> AnalysisResult:
    """An AnalysisResult with populated metrics matching sample_records."""
    ar = AnalysisResult()
    ar.total_files = 5
    ar.total_size_bytes = 5_032_572
    ar.by_category = {
        "Codigo": {
            "count": 3,
            "bytes": 2548,
            "percent": 0.05,
            "avg_size": 849.3,
            "newest": datetime(2025, 6, 15),
            "oldest": datetime(2025, 6, 15),
        },
        "Multimedia": {
            "count": 1,
            "bytes": 5_000_000,
            "percent": 99.35,
            "avg_size": 5_000_000.0,
            "newest": datetime(2025, 6, 15),
            "oldest": datetime(2025, 6, 15),
        },
        "Oficina": {
            "count": 1,
            "bytes": 30_000,
            "percent": 0.6,
            "avg_size": 30_000.0,
            "newest": datetime(2025, 6, 15),
            "oldest": datetime(2025, 6, 15),
        },
    }
    ar.timeline = {
        "2025-06-01": 3,
        "2024-12-01": 1,
        "2025-01-01": 1,
    }
    ar.top_largest = [
        {"path": "/tmp/b/photo.jpg", "size_bytes": 5_000_000, "category": "Multimedia", "mtime": datetime(2025, 6, 15)},
        {"path": "/tmp/a/report.xlsx", "size_bytes": 30_000, "category": "Oficina", "mtime": datetime(2025, 6, 15)},
    ]
    ar.inactive_files = [
        {"path": "/tmp/a/readme.md", "size_bytes": 2048, "category": "Codigo", "mtime": datetime(2025, 6, 15), "days_inactive": 200},
        {"path": "/tmp/a/script.py", "size_bytes": 500, "category": "Codigo", "mtime": datetime(2025, 6, 15), "days_inactive": 180},
    ]
    ar.zero_byte_files = [
        {"path": "/tmp/a/Makefile", "category": "Codigo", "mtime": datetime(2025, 6, 15)},
    ]
    ar.empty_directories = []
    ar.duplicates_by_name = {
        "config.yml": ["/tmp/a/config.yml", "/tmp/b/config.yml", "/tmp/c/config.yml"],
    }
    ar.permission_issues = [
        {"path": "/tmp/a/script.py", "permissions": "777", "issue": "777"},
    ]
    return ar


# ---------------------------------------------------------------------------
# 5.1 BaseReporter is abstract
# ---------------------------------------------------------------------------


class TestBaseReporterAbstract:
    def test_cannot_instantiate(self) -> None:
        with pytest.raises(TypeError):
            BaseReporter()  # type: ignore[abstract]


# ---------------------------------------------------------------------------
# 5.2 Generate with valid data
# ---------------------------------------------------------------------------


class TestGenerateValid:
    def test_creates_file_and_returns_path(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        reporter = ExcelReporter()
        result = reporter.generate(sample_records, sample_analysis, out)

        assert result == out
        assert out.exists()
        assert out.stat().st_size > 0

    def test_file_is_loadable(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        assert wb is not None
        wb.close()


# ---------------------------------------------------------------------------
# 5.3 Empty input
# ---------------------------------------------------------------------------


class TestEmptyInput:
    def test_empty_produces_valid_xlsx(self, tmp_path: Path) -> None:
        out = tmp_path / "empty.xlsx"
        ExcelReporter().generate([], AnalysisResult(), out)

        assert out.exists()
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 8
        wb.close()

    def test_empty_sheets_have_headers_only(self, tmp_path: Path) -> None:
        out = tmp_path / "empty.xlsx"
        ExcelReporter().generate([], AnalysisResult(), out)
        wb = load_workbook(out)

        # Inventario should have header row only
        inv = wb["Inventario Completo"]
        assert inv.max_row >= 1  # at least header
        # No data rows beyond possible header
        wb.close()


# ---------------------------------------------------------------------------
# 5.4 Invalid output path
# ---------------------------------------------------------------------------


class TestInvalidOutputPath:
    def test_missing_parent_raises(self, tmp_path: Path) -> None:
        out = tmp_path / "nonexistent_dir" / "report.xlsx"
        with pytest.raises(FileNotFoundError):
            ExcelReporter().generate([], AnalysisResult(), out)

    def test_no_partial_file(self, tmp_path: Path) -> None:
        out = tmp_path / "nonexistent_dir" / "report.xlsx"
        with pytest.raises(FileNotFoundError):
            ExcelReporter().generate([], AnalysisResult(), out)
        assert not out.exists()


# ---------------------------------------------------------------------------
# 5.5 Sheet names and order
# ---------------------------------------------------------------------------


class TestSheetNamesOrder:
    def test_sheet_names_match(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        assert wb.sheetnames == SHEET_NAMES
        wb.close()


# ---------------------------------------------------------------------------
# 5.6 Dashboard KPIs
# ---------------------------------------------------------------------------


class TestDashboard:
    @staticmethod
    def _find_kpi(ws, label: str):
        """Find a KPI row by label, return (row, value)."""
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == label:
                return row, ws.cell(row=row, column=2).value
        return None, None

    @staticmethod
    def _find_section(ws, title: str) -> int | None:
        """Find a section header row by title."""
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == title:
                return row
        return None

    def test_dashboard_title(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        assert wb["Dashboard"].cell(row=1, column=1).value == "Dashboard"
        wb.close()

    def test_kpi_health_score(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        row, val = self._find_kpi(wb["Dashboard"], "Health Score")
        assert row is not None, "Health Score KPI not found"
        assert "100" in str(val) or isinstance(val, (int, float))
        wb.close()

    def test_kpi_total_archivos(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        row, val = self._find_kpi(wb["Dashboard"], "Total Archivos")
        assert row is not None, "Total Archivos KPI not found"
        assert val == 5
        wb.close()

    def test_kpi_tamano_total(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        row, val = self._find_kpi(wb["Dashboard"], "Tamaño Total (MB)")
        assert row is not None
        assert isinstance(val, float)
        assert val == ExcelReporter._bytes_to_mb(5_032_572)
        wb.close()

    def test_kpi_alertas(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        row, val = self._find_kpi(wb["Dashboard"], "Alertas Activas")
        assert row is not None
        assert val == 5  # 1 zero-byte + 1 permission + 3 duplicates
        wb.close()

    def test_kpi_duplicados(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        row, val = self._find_kpi(wb["Dashboard"], "Archivos Duplicados")
        assert row is not None
        wb.close()

    def test_kpi_inactivos(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        row, val = self._find_kpi(wb["Dashboard"], "Archivos Inactivos")
        assert row is not None
        wb.close()

    def test_kpi_zero_bytes(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        row, val = self._find_kpi(wb["Dashboard"], "Archivos 0 bytes")
        assert row is not None
        assert val == 1
        wb.close()

    def test_kpi_tamano_promedio(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        row, val = self._find_kpi(wb["Dashboard"], "Tamaño Promedio (MB)")
        assert row is not None
        assert isinstance(val, float)
        wb.close()

    def test_kpi_directorios_vacios(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        row, val = self._find_kpi(wb["Dashboard"], "Directorios Vacíos")
        assert row is not None
        wb.close()

    def test_kpi_extension_mas_comun(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        row, val = self._find_kpi(wb["Dashboard"], "Extensión Más Común")
        assert row is not None
        assert isinstance(val, str)
        wb.close()

    def test_kpi_archivo_mas_pesado(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        row, val = self._find_kpi(wb["Dashboard"], "Archivo Más Pesado")
        assert row is not None
        assert "photo.jpg" in str(val)
        wb.close()

    def test_dashboard_top5_categories(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        ws = wb["Dashboard"]

        row = self._find_section(ws, "Top 5 Categorías por Tamaño")
        assert row is not None, "Top 5 Categories section not found"
        assert ws.cell(row=row + 1, column=1).value == "Categoría"
        assert ws.cell(row=row + 1, column=2).value == "Cantidad"
        assert ws.cell(row=row + 1, column=3).value == "Tamaño (MB)"
        # First category should be Multimedia (largest)
        assert ws.cell(row=row + 2, column=1).value == "Multimedia"
        wb.close()

    def test_dashboard_top5_directories(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        ws = wb["Dashboard"]

        row = self._find_section(ws, "Top 5 Directorios por Cantidad")
        assert row is not None, "Top 5 Directorios section not found"
        assert ws.cell(row=row + 1, column=1).value == "Directorio"
        assert ws.cell(row=row + 1, column=2).value == "Cantidad"
        wb.close()

    def test_dashboard_has_timeline_chart(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        ws = wb["Dashboard"]

        # Dashboard should have a chart embedded at E2
        assert len(ws._charts) >= 1, "No chart found on Dashboard sheet"
        wb.close()

    def test_dashboard_timeline_data_present(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        ws = wb["Dashboard"]

        # Timeline data is in col E — find the period values
        all_col_e = [ws.cell(row=r, column=5).value for r in range(1, ws.max_row + 1)]
        assert datetime(2024, 12, 1) in all_col_e
        assert datetime(2025, 1, 1) in all_col_e
        assert datetime(2025, 6, 1) in all_col_e
        wb.close()


# ---------------------------------------------------------------------------
# 5.7 Por Categoria rows
# ---------------------------------------------------------------------------


class TestPorCategoria:
    def test_row_count_matches_categories(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        ws = wb["Por Categoria"]

        # 3 categories + 1 header = 4 rows
        data_rows = ws.max_row - 1  # subtract header
        assert data_rows == 3
        wb.close()

    def test_size_columns_are_mb_floats(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        ws = wb["Por Categoria"]

        # Column C (Volumen MB) and E (Promedio MB) should be floats
        for row in range(2, ws.max_row + 1):
            vol = ws.cell(row=row, column=3).value
            avg = ws.cell(row=row, column=5).value
            assert isinstance(vol, (int, float)), f"Row {row} col C not numeric: {vol}"
            assert isinstance(avg, (int, float)), f"Row {row} col E not numeric: {avg}"

        wb.close()


# ---------------------------------------------------------------------------
# 5.8 Timeline sort order
# ---------------------------------------------------------------------------


class TestTimelineSort:
    def test_chronological_ascending(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        ws = wb["Timeline"]

        periods = []
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val:
                periods.append(val)

        assert periods == sorted(periods)
        assert periods == [datetime(2024, 12, 1), datetime(2025, 1, 1), datetime(2025, 6, 1)]
        wb.close()


# ---------------------------------------------------------------------------
# 5.8b Top Archivos Pesados has Nombre column with MB sizes
# ---------------------------------------------------------------------------


class TestTopPesadosNombre:
    def test_nombre_column_present(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        ws = wb["Top Archivos Pesados"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Nombre" in headers
        assert headers == ["Ruta", "Nombre", "Tamaño (MB)", "Categoría", "Última Modificación"]

        # Verify data rows derive Nombre from path
        assert ws.cell(row=2, column=2).value == "photo.jpg"
        assert ws.cell(row=3, column=2).value == "report.xlsx"

        # Verify size is numeric (int or float)
        assert isinstance(ws.cell(row=2, column=3).value, (int, float))
        wb.close()


# ---------------------------------------------------------------------------
# 5.8c Archivos Inactivos has Nombre column with MB sizes
# ---------------------------------------------------------------------------


class TestInactivosNombre:
    def test_nombre_column_present(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        ws = wb["Archivos Inactivos"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Nombre" in headers
        assert headers == ["Ruta", "Nombre", "Tamaño (MB)", "Categoría", "Última Modificación", "Días Inactivo"]

        # Verify data rows derive Nombre from path (sorted by days_inactive desc)
        assert ws.cell(row=2, column=2).value == "readme.md"
        assert ws.cell(row=3, column=2).value == "script.py"

        # Verify size is numeric (int or float — 0.0 may be stored as int 0)
        assert isinstance(ws.cell(row=2, column=3).value, (int, float))
        wb.close()


# ---------------------------------------------------------------------------
# 5.9 Alertas aggregates all sources
# ---------------------------------------------------------------------------


class TestAlertas:
    def test_aggregates_all_sources(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        ws = wb["Alertas"]

        data_rows = ws.max_row - 1  # subtract header
        # Expected: 1 zero-byte + 1 permission + 3 duplicates + 1 no-extension = 6
        assert data_rows >= 6

        # Check alert types present
        types = set()
        for row in range(2, ws.max_row + 1):
            types.add(ws.cell(row=row, column=1).value)

        assert "0 bytes" in types
        assert any("Permisos" in t for t in types if t)
        assert "Duplicado" in types
        assert "Sin extension" in types

        wb.close()


# ---------------------------------------------------------------------------
# 5.10 Inventario has autofilter
# ---------------------------------------------------------------------------


class TestInventarioAutofilter:
    def test_autofilter_set(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        ws = wb["Inventario Completo"]

        assert ws.auto_filter.ref is not None
        assert ws.auto_filter.ref != ""
        wb.close()


# ---------------------------------------------------------------------------
# 5.11 _format_bytes parametrized
# ---------------------------------------------------------------------------


class TestFormatBytes:
    @pytest.mark.parametrize(
        "size_bytes, expected",
        [
            (0, "0 B"),
            (512, "512 B"),
            (1023, "1023 B"),
            (1024, "1.00 KB"),
            (1536, "1.50 KB"),
            (1_048_576, "1.00 MB"),
            (1_073_741_824, "1.00 GB"),
            (2_500_000_000, "2.33 GB"),
        ],
    )
    def test_format_bytes(self, size_bytes: int, expected: str) -> None:
        assert ExcelReporter._format_bytes(size_bytes) == expected


# ---------------------------------------------------------------------------
# 5.12 Header styling
# ---------------------------------------------------------------------------


class TestHeaderStyling:
    def test_bold_headers_and_frozen_panes(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)

        for sheet_name in SHEET_NAMES:
            ws = wb[sheet_name]
            # Header row should be bold
            first_cell = ws.cell(row=1, column=1)
            assert first_cell.font.bold, f"Header not bold in {sheet_name}"
            # Frozen panes
            assert ws.freeze_panes == "A2", f"Freeze panes wrong in {sheet_name}"

        wb.close()


# ---------------------------------------------------------------------------
# 6.1 _bytes_to_mb parametrized
# ---------------------------------------------------------------------------


class TestBytesToMb:
    @pytest.mark.parametrize(
        "size_bytes, expected",
        [
            (0, 0.0),
            (1_048_576, 1.0),
            (2_097_152, 2.0),
            (500, 0.0),
            (5_000_000, 4.77),
            (10_485_760, 10.0),
        ],
    )
    def test_bytes_to_mb(self, size_bytes: int, expected: float) -> None:
        assert ExcelReporter._bytes_to_mb(size_bytes) == expected


# ---------------------------------------------------------------------------
# 6.3 Autofilter on all tabular sheets
# ---------------------------------------------------------------------------


class TestAutofilterAllSheets:
    """Verify autofilter is set on all 6 tabular sheets + Inventario."""

    AUTOFILTER_SHEETS = [
        "Por Categoria",
        "Timeline",
        "Top Archivos Pesados",
        "Archivos Inactivos",
        "Alertas",
        "Por Directorio",
        "Inventario Completo",
    ]

    def test_autofilter_present_on_tabular_sheets(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)

        for sheet_name in self.AUTOFILTER_SHEETS:
            ws = wb[sheet_name]
            assert ws.auto_filter.ref is not None, f"No autofilter on {sheet_name}"
            assert ws.auto_filter.ref != "", f"Empty autofilter on {sheet_name}"

        wb.close()

    def test_autofilter_on_empty_data(self, tmp_path: Path) -> None:
        """Empty sheets should still have autofilter covering headers."""
        out = tmp_path / "empty.xlsx"
        ExcelReporter().generate([], AnalysisResult(), out)
        wb = load_workbook(out)

        for sheet_name in self.AUTOFILTER_SHEETS:
            ws = wb[sheet_name]
            ref = ws.auto_filter.ref
            assert ref is not None, f"No autofilter on empty {sheet_name}"
            assert ref != "", f"Empty autofilter ref on empty {sheet_name}"
            # Should end with row 1 (header only)
            assert ref.endswith("1"), f"Autofilter ref {ref} on {sheet_name} should end at row 1 for empty data"

        wb.close()

    def test_dashboard_has_no_autofilter(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)

        ws = wb["Dashboard"]
        assert ws.auto_filter.ref is None or ws.auto_filter.ref == ""

        wb.close()


# ---------------------------------------------------------------------------
# 6.5 Timeline chart assertion
# ---------------------------------------------------------------------------


class TestTimelineChart:
    def test_chart_exists_with_data(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        """Timeline with >= 2 periods should have a LineChart."""
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        ws = wb["Timeline"]

        assert len(ws._charts) == 1
        chart = ws._charts[0]
        # After save/load, chart.title is a Title object with rich text runs
        title_obj = chart.title
        runs = title_obj.tx.rich.paragraphs[0].r
        title_text = "".join(run.t for run in runs)
        assert title_text == "Archivos por Período"

        wb.close()

    def test_no_chart_on_empty_timeline(self, tmp_path: Path) -> None:
        """Empty timeline should have no chart."""
        out = tmp_path / "empty.xlsx"
        ExcelReporter().generate([], AnalysisResult(), out)
        wb = load_workbook(out)
        ws = wb["Timeline"]

        assert len(ws._charts) == 0

        wb.close()

    def test_no_chart_on_single_period(self, tmp_path: Path) -> None:
        """Timeline with only 1 period should have no chart."""
        ar = AnalysisResult()
        ar.timeline = {"2025-01-01": 5}
        out = tmp_path / "single.xlsx"
        ExcelReporter().generate([], ar, out)
        wb = load_workbook(out)
        ws = wb["Timeline"]

        assert len(ws._charts) == 0

        wb.close()


# ---------------------------------------------------------------------------
# 6.6 Inventario uses MB values
# ---------------------------------------------------------------------------


class TestInventarioMB:
    def test_size_column_is_mb_float(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        ws = wb["Inventario Completo"]

        # Header check
        assert ws.cell(row=1, column=4).value == "Tamaño (MB)"

        # Data cells should be numeric floats
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=4).value
            assert isinstance(val, (int, float)), f"Row {row} size not numeric: {val}"

        wb.close()


# ---------------------------------------------------------------------------
# Task file-author-metadata: "Autor" column in Inventario Completo
# ---------------------------------------------------------------------------

class TestInventarioAutorColumn:
    """Tests for the Autor column in the Inventario Completo sheet."""

    def test_inventario_has_autor_header(
        self, tmp_path: Path, sample_records: list[FileRecord], sample_analysis: AnalysisResult
    ) -> None:
        """Column 13 (M) header is 'Autor'."""
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate(sample_records, sample_analysis, out)
        wb = load_workbook(out)
        ws = wb["Inventario Completo"]
        assert ws.cell(row=1, column=13).value == "Autor"
        wb.close()

    def test_inventario_author_value(
        self, tmp_path: Path, sample_analysis: AnalysisResult
    ) -> None:
        """A record with author='Alice' has 'Alice' in the Autor column."""
        from dataclasses import replace as dc_replace
        rec = _make_record("doc.docx", ".docx", 1000, "Oficina", "/tmp/a")
        rec_with_author = dc_replace(rec, author="Alice")
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate([rec_with_author], sample_analysis, out)
        wb = load_workbook(out)
        ws = wb["Inventario Completo"]
        assert ws.cell(row=2, column=13).value == "Alice"
        wb.close()

    def test_inventario_author_none_empty(
        self, tmp_path: Path, sample_analysis: AnalysisResult
    ) -> None:
        """A record with author=None has empty string in the Autor column."""
        rec = _make_record("file.txt", ".txt", 100, "Texto", "/tmp/a")
        # author is None by default
        out = tmp_path / "report.xlsx"
        ExcelReporter().generate([rec], sample_analysis, out)
        wb = load_workbook(out)
        ws = wb["Inventario Completo"]
        val = ws.cell(row=2, column=13).value
        assert val == "" or val is None
        wb.close()
