"""Tests for Excel overflow strategy — shard mode, CSV mode, and warnings."""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import Workbook

from fsaudit.cli import build_parser, main
from fsaudit.scanner.models import FileRecord


# ============================================================================
# T01: Constants
# ============================================================================


def test_excel_max_rows_constant():
    """EXCEL_MAX_ROWS must equal openpyxl's row limit (1,048,576)."""
    from fsaudit.reporter.excel_reporter import EXCEL_MAX_ROWS

    assert EXCEL_MAX_ROWS == 1_048_576


def test_max_inventory_rows_constant():
    """MAX_INVENTORY_ROWS is one less than EXCEL_MAX_ROWS to leave room for header."""
    from fsaudit.reporter.excel_reporter import MAX_INVENTORY_ROWS

    assert MAX_INVENTORY_ROWS == 1_048_575


# ============================================================================
# T02: AuditResult.overflow_warning field
# ============================================================================


def test_audit_result_overflow_warning_default_none():
    """AuditResult() must have overflow_warning defaulting to None."""
    from fsaudit.api import AuditResult

    result = AuditResult(
        records=[],
        analysis=None,  # type: ignore[arg-type]
        scan_result=None,  # type: ignore[arg-type]
    )
    assert result.overflow_warning is None


def test_audit_result_overflow_warning_can_be_set():
    """overflow_warning can be set to a string value."""
    from fsaudit.api import AuditResult

    result = AuditResult(
        records=[],
        analysis=None,  # type: ignore[arg-type]
        scan_result=None,  # type: ignore[arg-type]
        overflow_warning="Inventario dividido en 3 hojas (150 archivos totales). Estrategia: shard",
    )
    assert result.overflow_warning == (
        "Inventario dividido en 3 hojas (150 archivos totales). Estrategia: shard"
    )


# ============================================================================
# T03: ExcelReporter overflow_strategy constructor param
# ============================================================================


def test_excel_reporter_default_overflow_strategy():
    """ExcelReporter() defaults to overflow_strategy='shard'."""
    from fsaudit.reporter.excel_reporter import ExcelReporter

    reporter = ExcelReporter()
    assert reporter.overflow_strategy == "shard"


def test_excel_reporter_explicit_overflow_strategy():
    """ExcelReporter(overflow_strategy='csv') stores it."""
    from fsaudit.reporter.excel_reporter import ExcelReporter

    reporter = ExcelReporter(overflow_strategy="csv")
    assert reporter.overflow_strategy == "csv"


# ============================================================================
# T04: _inventory_sheet_names tests (RED)
# ============================================================================


class TestInventorySheetNames:
    """Tests for _inventory_sheet_names() with MAX_INVENTORY_ROWS=50."""

    def test_single_sheet_below_max(self, monkeypatch):
        """30 records → single 'Inventario Completo'."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 50
        )
        reporter = ExcelReporter()
        names = reporter._inventory_sheet_names(30)
        assert names == ["Inventario Completo"]

    def test_three_shards_above_max(self, monkeypatch):
        """120 records → 3 shards 'Inventario 1 de 3', 'Inventario 2 de 3', 'Inventario 3 de 3'."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 50
        )
        reporter = ExcelReporter()
        names = reporter._inventory_sheet_names(120)
        assert names == ["Inventario 1 de 3", "Inventario 2 de 3", "Inventario 3 de 3"]

    def test_exactly_max_records_single_sheet(self, monkeypatch):
        """Exactly 50 records → single 'Inventario Completo' (boundary)."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 50
        )
        reporter = ExcelReporter()
        names = reporter._inventory_sheet_names(50)
        assert names == ["Inventario Completo"]

    def test_just_above_max_two_shards(self, monkeypatch):
        """51 records → 2 shards 'Inventario 1 de 2', 'Inventario 2 de 2'."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 50
        )
        reporter = ExcelReporter()
        names = reporter._inventory_sheet_names(51)
        assert names == ["Inventario 1 de 2", "Inventario 2 de 2"]


# ============================================================================
# T06: _write_inventory_sheets tests (RED)
# ============================================================================


def _make_overflow_record(index: int) -> FileRecord:
    """Build a minimal FileRecord for overflow testing."""
    import datetime as dt

    return FileRecord(
        path=Path(f"/test/file_{index}.txt"),
        name=f"file_{index}.txt",
        extension=".txt",
        size_bytes=1024,
        mtime=dt.datetime(2025, 6, 15),
        creation_time=dt.datetime(2025, 1, 1),
        atime=dt.datetime(2025, 6, 20),
        depth=1,
        is_hidden=False,
        permissions="644",
        category="Codigo",
        parent_dir="/test",
    )


class TestWriteInventorySheets:
    """Tests for _write_inventory_sheets() with MAX_INVENTORY_ROWS=50."""

    def test_three_shards_created(self, monkeypatch):
        """120 records → 3 inventory sheets with correct names."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 50
        )
        reporter = ExcelReporter()
        records = [_make_overflow_record(i) for i in range(120)]
        wb = Workbook()
        # Remove the default sheet so we only have our inventory sheets
        wb.remove(wb.active)

        shard_count = reporter._write_inventory_sheets(wb, records)

        assert shard_count == 3
        assert wb.sheetnames == [
            "Inventario 1 de 3",
            "Inventario 2 de 3",
            "Inventario 3 de 3",
        ]

    def test_each_shard_has_correct_row_count(self, monkeypatch):
        """Each shard: ≤50 data rows + 1 header row."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 50
        )
        reporter = ExcelReporter()
        records = [_make_overflow_record(i) for i in range(120)]
        wb = Workbook()
        wb.remove(wb.active)

        reporter._write_inventory_sheets(wb, records)

        # Shard 1: 50 data rows + 1 header = 51 rows
        ws1 = wb["Inventario 1 de 3"]
        assert ws1.max_row == 51

        # Shard 2: 50 data rows + 1 header = 51 rows
        ws2 = wb["Inventario 2 de 3"]
        assert ws2.max_row == 51

        # Shard 3: 20 data rows + 1 header = 21 rows
        ws3 = wb["Inventario 3 de 3"]
        assert ws3.max_row == 21

    def test_each_shard_has_autofilter(self, monkeypatch):
        """Each shard must have autofilter set."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 50
        )
        reporter = ExcelReporter()
        records = [_make_overflow_record(i) for i in range(120)]
        wb = Workbook()
        wb.remove(wb.active)

        reporter._write_inventory_sheets(wb, records)

        for name in wb.sheetnames:
            ws = wb[name]
            assert ws.auto_filter.ref is not None, (
                f"No autofilter on {name}"
            )
            assert ws.auto_filter.ref != "", f"Empty autofilter on {name}"

    def test_each_shard_has_freeze_panes_a2(self, monkeypatch):
        """Each shard must have freeze panes at A2."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 50
        )
        reporter = ExcelReporter()
        records = [_make_overflow_record(i) for i in range(120)]
        wb = Workbook()
        wb.remove(wb.active)

        reporter._write_inventory_sheets(wb, records)

        for name in wb.sheetnames:
            ws = wb[name]
            assert ws.freeze_panes == "A2", (
                f"Freeze panes not at A2 on {name}: {ws.freeze_panes}"
            )


# ============================================================================
# T08: _write_inventario_chunk tests (RED)
# ============================================================================


class TestWriteInventarioChunk:
    """Tests for _write_inventario_chunk() — per-shard output structure."""

    def test_chunk_headers_match_full_inventario(self, monkeypatch):
        """Column headers in chunk must match the original Inventario headers."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 50
        )
        reporter = ExcelReporter()
        records = [_make_overflow_record(i) for i in range(10)]

        wb = Workbook()
        ws = wb.active
        ws.title = "Test Chunk"

        reporter._write_inventario_chunk(ws, records)

        expected_headers = [
            "Ruta", "Nombre", "Extensión", "Tamaño (MB)",
            "Categoría", "Fecha Modificación", "Fecha Creación",
            "Último Acceso", "Profundidad", "Oculto",
            "Permisos", "Directorio Padre", "Autor",
        ]
        for col, expected in enumerate(expected_headers, start=1):
            assert ws.cell(row=1, column=col).value == expected, (
                f"Column {col}: expected '{expected}'"
            )

    def test_chunk_autofilter_covers_only_chunk_rows(self, monkeypatch):
        """Autofilter range must cover exactly the chunk rows + header."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 50
        )
        reporter = ExcelReporter()
        records = [_make_overflow_record(i) for i in range(7)]

        wb = Workbook()
        ws = wb.active
        ws.title = "Test Chunk"

        reporter._write_inventario_chunk(ws, records)

        assert ws.auto_filter.ref is not None
        # 7 data rows + 1 header = row 8, 13 columns = column M
        assert ws.auto_filter.ref == "A1:M8", (
            f"Expected A1:M8, got {ws.auto_filter.ref}"
        )

    def test_chunk_freeze_panes_at_a2(self, monkeypatch):
        """Freeze panes must be at A2."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 50
        )
        reporter = ExcelReporter()
        records = [_make_overflow_record(i) for i in range(5)]

        wb = Workbook()
        ws = wb.active
        ws.title = "Test Chunk"

        reporter._write_inventario_chunk(ws, records)

        assert ws.freeze_panes == "A2"

    def test_chunk_data_values_correct(self, monkeypatch):
        """Data rows must contain correct values from records."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 50
        )
        reporter = ExcelReporter()
        records = [_make_overflow_record(i) for i in range(3)]

        wb = Workbook()
        ws = wb.active
        ws.title = "Test Chunk"

        reporter._write_inventario_chunk(ws, records)

        # Row 2 (first data row): file_0.txt
        assert ws.cell(row=2, column=2).value == "file_0.txt"
        assert ws.cell(row=2, column=5).value == "Codigo"
        # Size column should be a float (MB)
        assert isinstance(ws.cell(row=2, column=4).value, float)

        # Row 4 (third data row): file_2.txt
        assert ws.cell(row=4, column=2).value == "file_2.txt"


# ============================================================================
# T11: Dashboard overflow rows tests (RED)
# ============================================================================


class TestDashboardOverflowRows:
    """Tests for overflow warning rows in the Dashboard sheet."""

    def _generate_and_get_dashboard(self, monkeypatch, n_records: int):
        """Helper: generate report with n_records and return Dashboard ws."""
        from fsaudit.analyzer.metrics import AnalysisResult
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 50
        )

        records = [_make_overflow_record(i) for i in range(n_records)]
        analysis = AnalysisResult()
        analysis.total_files = n_records
        analysis.total_size_bytes = n_records * 1024

        reporter = ExcelReporter()

        from openpyxl import Workbook
        wb = Workbook()
        # Create fixed sheets
        from fsaudit.reporter.excel_reporter import SHEET_NAMES
        for idx, name in enumerate(SHEET_NAMES):
            if idx == 0:
                wb.active.title = name
            else:
                wb.create_sheet(title=name)
        # Create inventory sheets
        inventory_count = reporter._write_inventory_sheets(wb, records)
        overflow = inventory_count > 1
        if overflow:
            reporter._overflow_warning = (
                f"Inventario dividido en {inventory_count} hojas "
                f"({n_records} archivos totales). Estrategia: shard"
            )

        reporter._write_dashboard(
            wb["Dashboard"], analysis, records,
            overflow=overflow,
            inventory_count=inventory_count,
        )
        return wb["Dashboard"]

    def test_overflow_rows_present(self, monkeypatch):
        """120 records → rows 2-3 contain overflow message + strategy."""
        ws = self._generate_and_get_dashboard(monkeypatch, 120)

        # Row 2: Overflow warning
        assert ws.cell(row=2, column=1).value == "⚠ OVERFLOW"
        assert "120" in str(ws.cell(row=2, column=2).value)
        assert "3 hojas" in str(ws.cell(row=2, column=2).value)

        # Row 3: Strategy
        assert ws.cell(row=3, column=1).value == "Estrategia"
        assert ws.cell(row=3, column=2).value == "shard"

    def test_no_overflow_rows_when_below_max(self, monkeypatch):
        """30 records → no overflow rows; KPIs start at row 2."""
        ws = self._generate_and_get_dashboard(monkeypatch, 30)

        # Row 2 should be a KPI, not overflow
        row2_label = ws.cell(row=2, column=1).value
        assert row2_label != "⚠ OVERFLOW", "Row 2 should NOT be overflow"
        assert row2_label == "Health Score", f"Expected 'Health Score', got '{row2_label}'"

    def test_overflow_shifts_kpis_down(self, monkeypatch):
        """Overflow pushes KPIs down by 2 rows."""
        ws = self._generate_and_get_dashboard(monkeypatch, 120)

        # Health Score should now be at row 4 (originally row 2)
        assert ws.cell(row=4, column=1).value == "Health Score"

    def test_overflow_warning_not_set_when_below_max(self, monkeypatch):
        """_overflow_warning should be None when no overflow."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 50
        )
        reporter = ExcelReporter()
        records = [_make_overflow_record(i) for i in range(30)]
        assert len(records) <= 50
        # No overflow means _overflow_warning doesn't exist yet
        assert not hasattr(reporter, "_overflow_warning") or reporter._overflow_warning is None


# ============================================================================
# T13: Backward compatibility test
# ============================================================================


class TestBackwardCompatibility:
    """Ensure sub-limit files produce identical output to v0.10.0."""

    def test_sub_limit_single_inventario_completo(self, monkeypatch, tmp_path):
        """30 records → single 'Inventario Completo' sheet, 8 sheets total, no warning."""
        from fsaudit.analyzer.metrics import AnalysisResult
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 50
        )
        reporter = ExcelReporter(overflow_strategy="shard")
        records = [_make_overflow_record(i) for i in range(30)]
        analysis = AnalysisResult()
        analysis.total_files = 30
        analysis.total_size_bytes = 30 * 1024

        output = tmp_path / "test_output.xlsx"
        reporter.generate(records, analysis, output)

        from openpyxl import load_workbook
        wb = load_workbook(str(output))

        # Should have 7 fixed sheets + 1 "Inventario Completo" = 8 total
        assert len(wb.sheetnames) == 8
        assert "Inventario Completo" in wb.sheetnames
        # No overflow warning
        assert reporter._overflow_warning is None


# ============================================================================
# T14: CSV output tests — BOM, columns, no xlsx
# ============================================================================


class TestCSVOutput:
    """Tests for CSV mode output format."""

    def test_csv_file_created(self, tmp_path):
        """CSV mode creates a .csv file and NO .xlsx file."""
        from fsaudit.analyzer.metrics import AnalysisResult
        from fsaudit.reporter.excel_reporter import ExcelReporter

        reporter = ExcelReporter(overflow_strategy="csv")
        records = [_make_overflow_record(i) for i in range(10)]
        analysis = AnalysisResult()
        analysis.total_files = 10
        analysis.total_size_bytes = 10 * 1024

        output_xlsx = tmp_path / "testdir_audit_2025-06-15.xlsx"
        result_path = reporter.generate(records, analysis, output_xlsx)

        # Should return a .csv path
        assert result_path.suffix == ".csv"
        assert result_path.exists()

        # Should NOT create .xlsx file
        assert not output_xlsx.exists()

    def test_csv_has_utf8_bom(self, tmp_path):
        """CSV file must start with UTF-8 BOM bytes."""
        from fsaudit.analyzer.metrics import AnalysisResult
        from fsaudit.reporter.excel_reporter import ExcelReporter

        reporter = ExcelReporter(overflow_strategy="csv")
        records = [_make_overflow_record(i) for i in range(10)]
        analysis = AnalysisResult()
        analysis.total_files = 10
        analysis.total_size_bytes = 10 * 1024

        output_xlsx = tmp_path / "testdir_audit_2025-06-15.xlsx"
        result_path = reporter.generate(records, analysis, output_xlsx)

        raw = result_path.read_bytes()
        assert raw[:3] == b"\xef\xbb\xbf", "CSV must start with UTF-8 BOM"

    def test_csv_has_13_columns(self, tmp_path):
        """CSV header row must have exactly 13 columns matching Inventario headers."""
        import csv as csv_mod

        from fsaudit.analyzer.metrics import AnalysisResult
        from fsaudit.reporter.excel_reporter import ExcelReporter

        reporter = ExcelReporter(overflow_strategy="csv")
        records = [_make_overflow_record(i) for i in range(10)]
        analysis = AnalysisResult()
        analysis.total_files = 10
        analysis.total_size_bytes = 10 * 1024

        output_xlsx = tmp_path / "testdir_audit_2025-06-15.xlsx"
        result_path = reporter.generate(records, analysis, output_xlsx)

        with open(result_path, encoding="utf-8-sig") as f:
            reader = csv_mod.reader(f)
            headers = next(reader)

        assert len(headers) == 13
        assert headers == [
            "Ruta", "Nombre", "Extensión", "Tamaño (MB)",
            "Categoría", "Fecha Modificación", "Fecha Creación",
            "Último Acceso", "Profundidad", "Oculto",
            "Permisos", "Directorio Padre", "Autor",
        ]

    def test_csv_contains_all_records(self, tmp_path):
        """CSV data rows must contain all input records."""
        import csv as csv_mod

        from fsaudit.analyzer.metrics import AnalysisResult
        from fsaudit.reporter.excel_reporter import ExcelReporter

        reporter = ExcelReporter(overflow_strategy="csv")
        n = 100
        records = [_make_overflow_record(i) for i in range(n)]
        analysis = AnalysisResult()
        analysis.total_files = n
        analysis.total_size_bytes = n * 1024

        output_xlsx = tmp_path / "testdir_audit_2025-06-15.xlsx"
        result_path = reporter.generate(records, analysis, output_xlsx)

        with open(result_path, encoding="utf-8-sig") as f:
            reader = csv_mod.reader(f)
            rows = list(reader)

        # 1 header + n data rows
        assert len(rows) == n + 1

    def test_csv_sub_limit_also_produces_csv(self, tmp_path):
        """CSV mode ALWAYS produces CSV, even for sub-limit record counts (SC-04)."""
        from fsaudit.analyzer.metrics import AnalysisResult
        from fsaudit.reporter.excel_reporter import ExcelReporter

        reporter = ExcelReporter(overflow_strategy="csv")
        records = [_make_overflow_record(i) for i in range(5)]
        analysis = AnalysisResult()
        analysis.total_files = 5
        analysis.total_size_bytes = 5 * 1024

        output_xlsx = tmp_path / "testdir_audit_2025-06-15.xlsx"
        result_path = reporter.generate(records, analysis, output_xlsx)

        assert result_path.suffix == ".csv"
        assert result_path.exists()
        # No xlsx should exist
        assert not output_xlsx.exists()


# ============================================================================
# T15: CSV filename format test
# ============================================================================


class TestCSVFilenameFormat:
    """Tests for CSV output filename convention."""

    def test_csv_filename_format(self, tmp_path):
        """CSV filename must follow {folder}_inventory_{YYYY-MM-DD}.csv."""
        from fsaudit.analyzer.metrics import AnalysisResult
        from fsaudit.reporter.excel_reporter import ExcelReporter

        reporter = ExcelReporter(overflow_strategy="csv")
        records = [_make_overflow_record(i) for i in range(10)]
        analysis = AnalysisResult()
        analysis.total_files = 10
        analysis.total_size_bytes = 10 * 1024

        # The output_path uses _audit_ in stem, but CSV should use _inventory_
        output_xlsx = tmp_path / "mydir_audit_2025-06-15.xlsx"
        result_path = reporter.generate(records, analysis, output_xlsx)

        assert result_path.name == "mydir_inventory_2025-06-15.csv"

    def test_csv_overrides_xlsx_extension(self, tmp_path):
        """CSV mode must change .xlsx extension to .csv."""
        from fsaudit.analyzer.metrics import AnalysisResult
        from fsaudit.reporter.excel_reporter import ExcelReporter

        reporter = ExcelReporter(overflow_strategy="csv")
        records = [_make_overflow_record(i) for i in range(5)]
        analysis = AnalysisResult()
        analysis.total_files = 5
        analysis.total_size_bytes = 5 * 1024

        output_xlsx = tmp_path / "testdata_audit_2025-01-10.xlsx"
        result_path = reporter.generate(records, analysis, output_xlsx)

        assert result_path.suffix == ".csv"
        assert "_inventory_" in result_path.stem
        assert "_audit_" not in result_path.stem


# ============================================================================
# T18-T19: API parameter threading tests
# ============================================================================


class TestApiOverflowStrategy:
    """Tests for api.audit() overflow_strategy parameter (T18-T19)."""

    def test_api_audit_accepts_overflow_strategy(self, tmp_path):
        """api.audit() must accept overflow_strategy parameter."""
        from fsaudit.api import audit
        # Create a small directory to scan
        (tmp_path / "testfile.txt").write_text("hello")
        result = audit(str(tmp_path), overflow_strategy="csv", format=None)
        assert result is not None

    def test_api_audit_default_overflow_strategy_is_shard(self, tmp_path):
        """api.audit() defaults to overflow_strategy='shard'."""
        from fsaudit.api import audit
        (tmp_path / "testfile.txt").write_text("hello")
        # format=None means no report generated — just test the param is accepted
        result = audit(str(tmp_path), format=None)
        assert result is not None

    def test_api_audit_overflow_warning_propagation(self, tmp_path, monkeypatch):
        """api.audit() must propagate overflow_warning from ExcelReporter."""
        from fsaudit.api import audit
        from fsaudit.reporter.excel_reporter import MAX_INVENTORY_ROWS

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 5
        )
        # Create enough files to trigger overflow
        for i in range(12):
            (tmp_path / f"file_{i}.txt").write_text("x")

        result = audit(str(tmp_path), overflow_strategy="shard")
        assert result.overflow_warning is not None
        assert "hojas" in result.overflow_warning

    def test_api_audit_csv_strategy_no_warning_when_sub_limit(self, tmp_path):
        """api.audit() with csv strategy should set overflow_warning even with sub-limit files."""
        from fsaudit.api import audit

        (tmp_path / "testfile.txt").write_text("hello")
        result = audit(str(tmp_path), overflow_strategy="csv")
        # CSV mode always warns
        assert result.overflow_warning is not None


# ============================================================================
# T20: BaseReporter.generate() overflow_strategy kwarg
# ============================================================================


class TestBaseReporterOverflowStrategy:
    """Test that BaseReporter.generate() accepts overflow_strategy kwarg."""

    def test_base_reporter_generate_signature_accepts_overflow_strategy(self):
        """BaseReporter.generate() must accept overflow_strategy keyword arg."""
        from fsaudit.reporter.base import BaseReporter
        import inspect

        sig = inspect.signature(BaseReporter.generate)
        params = list(sig.parameters.keys())
        # Must have overflow_strategy as optional kwarg
        assert "overflow_strategy" in params or any(
            p.kind == inspect.Parameter.VAR_KEYWORD
            for p in sig.parameters.values()
        ), f"BaseReporter.generate() must accept overflow_strategy kwarg"


# ============================================================================
# T21-T22: CLI argument tests
# ============================================================================


class TestCliOverflowStrategy:
    """Tests for --overflow-strategy CLI flag (T21-T22)."""

    def test_cli_overflow_strategy_default_shard(self):
        """--overflow-strategy defaults to 'shard'."""
        parser = build_parser()
        args = parser.parse_args(["--path", "/tmp"])
        assert args.overflow_strategy == "shard"

    def test_cli_overflow_strategy_csv(self):
        """--overflow-strategy csv is accepted."""
        parser = build_parser()
        args = parser.parse_args(["--path", "/tmp", "--overflow-strategy", "csv"])
        assert args.overflow_strategy == "csv"

    def test_cli_overflow_strategy_shard(self):
        """--overflow-strategy shard is accepted."""
        parser = build_parser()
        args = parser.parse_args(["--path", "/tmp", "--overflow-strategy", "shard"])
        assert args.overflow_strategy == "shard"

    def test_cli_overflow_strategy_invalid_rejected(self):
        """--overflow-strategy with invalid value must cause SystemExit."""
        parser = build_parser()
        with pytest.raises(SystemExit):
            parser.parse_args(["--path", "/tmp", "--overflow-strategy", "sqlite"])

    def test_cli_overflow_strategy_passed_to_reporter(self, tmp_path):
        """main() must pass overflow_strategy to ExcelReporter."""
        from unittest.mock import patch, MagicMock
        from fsaudit.analyzer.metrics import AnalysisResult

        dummy_result = AnalysisResult()

        with patch("fsaudit.reporter.excel_reporter.ExcelReporter") as mock_reporter_cls, \
             patch("fsaudit.pipeline._analyze", return_value=dummy_result):
            mock_reporter_cls.return_value.generate.return_value = None
            result = main([
                "--path", str(tmp_path),
                "--output-dir", str(tmp_path),
                "--overflow-strategy", "csv",
            ])

        # Verify ExcelReporter was called with overflow_strategy="csv"
        mock_reporter_cls.assert_called_once_with(overflow_strategy="csv")


# ============================================================================
# T23: CLI Rich Panel warning on overflow
# ============================================================================


class TestCliOverflowWarning:
    """Tests for CLI warning Panel on overflow (T23)."""

    def test_cli_prints_overflow_panel(self, tmp_path, monkeypatch):
        """When overflow occurs, CLI must print a Rich Panel with overflow info."""
        from io import StringIO
        from unittest.mock import patch, MagicMock
        from rich.console import Console

        from fsaudit.analyzer.metrics import AnalysisResult
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 5
        )

        # Create enough files to trigger overflow
        for i in range(12):
            (tmp_path / f"file_{i}.txt").write_text("x")

        buf = StringIO()
        console = Console(file=buf, highlight=False)

        result = main(
            ["--path", str(tmp_path), "--output-dir", str(tmp_path)],
            _console=console,
        )
        assert result == 0
        output = buf.getvalue()
        # Should contain overflow warning text
        assert "OVERFLOW" in output or "hojas" in output or "overflow" in output.lower()

    def test_cli_no_overflow_panel_when_sub_limit(self, tmp_path):
        """When no overflow and shard mode, no warning panel should appear."""
        from io import StringIO
        from unittest.mock import patch
        from rich.console import Console

        (tmp_path / "testfile.txt").write_text("hello")

        buf = StringIO()
        console = Console(file=buf, highlight=False)

        result = main(
            ["--path", str(tmp_path), "--output-dir", str(tmp_path)],
            _console=console,
        )
        assert result == 0
        output = buf.getvalue()
        assert "OVERFLOW" not in output


# ============================================================================
# T24: ScanConfig overflow_strategy field
# ============================================================================


class TestScanConfigOverflowStrategy:
    """Test ScanConfig.overflow_strategy field (T24)."""

    @pytest.fixture(autouse=True)
    def _ensure_textual(self):
        """Skip test if textual is not installed."""
        pytest.importorskip("textual")

    def test_scan_config_has_overflow_strategy_default_shard(self):
        """ScanConfig must have overflow_strategy field defaulting to 'shard'."""
        from fsaudit.tui.models import ScanConfig

        cfg = ScanConfig(root=Path("/tmp"))
        assert cfg.overflow_strategy == "shard"

    def test_scan_config_overflow_strategy_can_be_csv(self):
        """ScanConfig.overflow_strategy can be set to 'csv'."""
        from fsaudit.tui.models import ScanConfig

        cfg = ScanConfig(root=Path("/tmp"), overflow_strategy="csv")
        assert cfg.overflow_strategy == "csv"


# ============================================================================
# T28: End-to-end CLI test — shard mode with patched small MAX
# ============================================================================


class TestE2EOverflowShard:
    """End-to-end CLI test for --overflow-strategy shard (T28)."""

    def test_e2e_shard_creates_multiple_sheets(self, tmp_path, monkeypatch):
        """CLI with --overflow-strategy shard and patched MAX must create sharded sheets."""
        from fsaudit.analyzer.metrics import AnalysisResult
        from fsaudit.reporter.excel_reporter import ExcelReporter
        from openpyxl import load_workbook

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 5
        )

        # Create enough files to trigger overflow
        for i in range(12):
            (tmp_path / f"file_{i}.txt").write_text("x" * (i + 1))

        monkeypatch.setattr(
            "fsaudit.cli.ExcelReporter", ExcelReporter
        )

        result = main([
            "--path", str(tmp_path),
            "--output-dir", str(tmp_path),
            "--overflow-strategy", "shard",
        ])
        assert result == 0

        # Find the generated xlsx file
        xlsx_files = list(tmp_path.glob("*.xlsx"))
        assert len(xlsx_files) == 1

        wb = load_workbook(str(xlsx_files[0]))
        # Should have multiple "Inventario" sheets
        inv_sheets = [s for s in wb.sheetnames if s.startswith("Inventario")]
        assert len(inv_sheets) > 1, f"Expected multiple inventory sheets, got {inv_sheets}"

    def test_e2e_shard_prints_overflow_warning(self, tmp_path, monkeypatch):
        """CLI with shard overflow must print warning Panel."""
        from io import StringIO
        from rich.console import Console

        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 5
        )

        for i in range(12):
            (tmp_path / f"file_{i}.txt").write_text("x" * (i + 1))

        monkeypatch.setattr("fsaudit.cli.ExcelReporter", ExcelReporter)

        buf = StringIO()
        console = Console(file=buf, highlight=False)
        result = main([
            "--path", str(tmp_path),
            "--output-dir", str(tmp_path),
            "--overflow-strategy", "shard",
        ], _console=console)

        assert result == 0
        output = buf.getvalue()
        # Panel should contain overflow-related text
        assert "1 de" in output or "overflow" in output.lower() or "hojas" in output.lower()


# ============================================================================
# T29: End-to-end CLI test — CSV mode with patched small MAX
# ============================================================================


class TestE2EOverflowCsv:
    """End-to-end CLI test for --overflow-strategy csv (T29)."""

    def test_e2e_csv_creates_csv_no_xlsx(self, tmp_path, monkeypatch):
        """CLI with --overflow-strategy csv must create CSV file, no xlsx."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 5
        )

        for i in range(12):
            (tmp_path / f"file_{i}.txt").write_text("x" * (i + 1))

        monkeypatch.setattr("fsaudit.cli.ExcelReporter", ExcelReporter)

        result = main([
            "--path", str(tmp_path),
            "--output-dir", str(tmp_path),
            "--overflow-strategy", "csv",
        ])
        assert result == 0

        # Should have a CSV file, no XLSX
        csv_files = list(tmp_path.glob("*.csv"))
        xlsx_files = list(tmp_path.glob("*.xlsx"))
        assert len(csv_files) == 1
        assert len(xlsx_files) == 0

    def test_e2e_csv_has_bom_and_correct_columns(self, tmp_path, monkeypatch):
        """CLI with --overflow-strategy csv must produce CSV with BOM and 13 columns."""
        import csv as csv_mod

        from fsaudit.reporter.excel_reporter import ExcelReporter

        monkeypatch.setattr(
            "fsaudit.reporter.excel_reporter.MAX_INVENTORY_ROWS", 5
        )

        for i in range(12):
            (tmp_path / f"file_{i}.txt").write_text("x" * (i + 1))

        monkeypatch.setattr("fsaudit.cli.ExcelReporter", ExcelReporter)

        result = main([
            "--path", str(tmp_path),
            "--output-dir", str(tmp_path),
            "--overflow-strategy", "csv",
        ])
        assert result == 0

        csv_files = list(tmp_path.glob("*.csv"))
        assert len(csv_files) == 1

        csv_path = csv_files[0]
        raw = csv_path.read_bytes()
        assert raw[:3] == b"\xef\xbb\xbf", "CSV must start with UTF-8 BOM"

        with open(csv_path, encoding="utf-8-sig") as f:
            reader = csv_mod.reader(f)
            headers = next(reader)

        assert len(headers) == 13
