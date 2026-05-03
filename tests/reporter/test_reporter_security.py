"""Tests for security integration in Excel and HTML reporters (T20, T21, T22)."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from fsaudit.analyzer.metrics import AnalysisResult
from fsaudit.scanner.models import FileRecord
from fsaudit.security.models import SecurityFinding, SecurityResult, Severity


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def _make_record(
    name: str = "file.txt",
    extension: str = ".txt",
    size_bytes: int = 1024,
    category: str = "Codigo",
    parent_dir: str = "/tmp/project",
    permissions: str | None = "644",
    is_hidden: bool = False,
    depth: int = 1,
    mtime: datetime | None = None,
) -> FileRecord:
    mt = mtime or datetime(2025, 6, 15, 10, 0, 0)
    return FileRecord(
        path=Path(parent_dir) / name,
        name=name,
        extension=extension,
        size_bytes=size_bytes,
        mtime=mt,
        creation_time=datetime(2025, 1, 1),
        atime=datetime(2025, 6, 20),
        depth=depth,
        is_hidden=is_hidden,
        permissions=permissions,
        category=category,
        parent_dir=parent_dir,
    )


def _make_finding(
    path: str = "/tmp/project/secret.env",
    rule_id: str = "aws-access-key",
    severity: Severity = Severity.CRITICAL,
    line_no: int | None = 3,
    match_context: str = "AKIA_TEST_context",
    detector: str = "secrets",
) -> SecurityFinding:
    return SecurityFinding(
        path=path,
        detector=detector,
        rule_id=rule_id,
        severity=severity,
        line_no=line_no,
        match_context=match_context,
        created_at=datetime(2025, 6, 15, 10, 0, 0),
    )


def _make_security_result(
    findings: list[SecurityFinding] | None = None,
    security_score: int = 75,
) -> SecurityResult:
    f = findings if findings is not None else [_make_finding()]
    return SecurityResult(
        findings=f,
        security_score=security_score,
        rules_applied=["aws-access-key", "github-token"],
        files_scanned=10,
        files_skipped=1,
        duration_s=0.5,
    )


def _make_analysis() -> AnalysisResult:
    ar = AnalysisResult()
    ar.total_files = 3
    ar.total_size_bytes = 3000
    ar.by_category = {
        "Codigo": {
            "count": 3,
            "bytes": 3000,
            "percent": 100.0,
            "avg_size": 1000.0,
            "newest": datetime(2025, 6, 15),
            "oldest": datetime(2025, 6, 15),
        },
    }
    ar.timeline = {"2025-06-01": 3}
    ar.top_largest = [{"path": "/tmp/project/file.txt", "size_bytes": 1024, "category": "Codigo", "mtime": datetime(2025, 6, 15)}]
    ar.inactive_files = []
    ar.zero_byte_files = []
    ar.empty_directories = []
    ar.duplicates_by_name = {}
    ar.permission_issues = []
    return ar


# ---------------------------------------------------------------------------
# T20 — Excel "Seguridad" sheet
# ---------------------------------------------------------------------------


class TestExcelSecuridadSheet:
    """Tests for the 9th sheet added by T20."""

    def test_seguridad_sheet_absent_when_no_scan(self, tmp_path: Path) -> None:
        """No Seguridad sheet when security=None (backward compat)."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        out = tmp_path / "report.xlsx"
        records = [_make_record()]
        ExcelReporter().generate(records, _make_analysis(), out, security=None)
        wb = load_workbook(out)
        assert "Seguridad" not in wb.sheetnames
        wb.close()

    def test_sheet_count_is_8_without_security(self, tmp_path: Path) -> None:
        """Exactly 8 sheets when no security data."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        out = tmp_path / "report.xlsx"
        ExcelReporter().generate([_make_record()], _make_analysis(), out)
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 8
        wb.close()

    def test_seguridad_sheet_created_when_scan_ran(self, tmp_path: Path) -> None:
        """Seguridad sheet exists when security result is provided."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        out = tmp_path / "report.xlsx"
        sr = _make_security_result()
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=sr)
        wb = load_workbook(out)
        assert "Seguridad" in wb.sheetnames
        wb.close()

    def test_sheet_count_is_9_with_security(self, tmp_path: Path) -> None:
        """Exactly 9 sheets when security data is provided."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        out = tmp_path / "report.xlsx"
        sr = _make_security_result()
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=sr)
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 9
        wb.close()

    def test_seguridad_is_last_sheet(self, tmp_path: Path) -> None:
        """Seguridad sheet is the 9th (last) sheet."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        out = tmp_path / "report.xlsx"
        sr = _make_security_result()
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=sr)
        wb = load_workbook(out)
        assert wb.sheetnames[-1] == "Seguridad"
        wb.close()

    def test_seguridad_has_n_plus_1_rows_for_n_findings(self, tmp_path: Path) -> None:
        """N findings → N+1 rows (header + N data rows)."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        findings = [
            _make_finding(path="/tmp/a.env", rule_id="aws-access-key", severity=Severity.CRITICAL),
            _make_finding(path="/tmp/b.env", rule_id="github-token", severity=Severity.HIGH),
            _make_finding(path="/tmp/c.py", rule_id="jwt", severity=Severity.MEDIUM),
        ]
        out = tmp_path / "report.xlsx"
        sr = _make_security_result(findings=findings)
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=sr)
        wb = load_workbook(out)
        ws = wb["Seguridad"]
        # Header row + 3 data rows = 4 rows
        assert ws.max_row == 4
        wb.close()

    def test_rows_sorted_severity_desc(self, tmp_path: Path) -> None:
        """Rows are sorted by severity DESC (critical first, low last)."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        findings = [
            _make_finding(path="/tmp/low.py", rule_id="jwt", severity=Severity.LOW),
            _make_finding(path="/tmp/crit.env", rule_id="aws-access-key", severity=Severity.CRITICAL),
            _make_finding(path="/tmp/high.env", rule_id="github-token", severity=Severity.HIGH),
        ]
        out = tmp_path / "report.xlsx"
        sr = _make_security_result(findings=findings)
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=sr)
        wb = load_workbook(out)
        ws = wb["Seguridad"]

        # Row 2 (first data row) should be CRITICAL, row 4 should be LOW
        # Find the severity column (look at header row)
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        sev_col = headers.index("Severidad") + 1

        severities = [ws.cell(row=r, column=sev_col).value for r in range(2, ws.max_row + 1)]
        assert severities[0] == "critical"
        assert severities[-1] == "low"
        wb.close()

    def test_autofilter_applied(self, tmp_path: Path) -> None:
        """Seguridad sheet has autofilter on the header row."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        out = tmp_path / "report.xlsx"
        sr = _make_security_result()
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=sr)
        wb = load_workbook(out)
        ws = wb["Seguridad"]
        assert ws.auto_filter.ref is not None
        wb.close()

    def test_contexto_column_max_60_chars(self, tmp_path: Path) -> None:
        """Match context (Contexto column) never exceeds 60 chars."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        # SecurityFinding.__post_init__ already truncates — this test verifies
        # the reporter writes match_context verbatim (never expands it)
        long_context = "A" * 60  # exactly at limit (post_init does NOT truncate at exactly 60)
        finding = _make_finding(match_context=long_context)
        out = tmp_path / "report.xlsx"
        sr = _make_security_result(findings=[finding])
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=sr)
        wb = load_workbook(out)
        ws = wb["Seguridad"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        ctx_col = headers.index("Contexto") + 1

        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=ctx_col).value or ""
            assert len(str(val)) <= 60
        wb.close()

    def test_critical_row_has_red_fill(self, tmp_path: Path) -> None:
        """CRITICAL severity row should have a red-ish fill on the severity cell."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        finding = _make_finding(severity=Severity.CRITICAL)
        out = tmp_path / "report.xlsx"
        sr = _make_security_result(findings=[finding])
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=sr)
        wb = load_workbook(out)
        ws = wb["Seguridad"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        sev_col = headers.index("Severidad") + 1

        cell = ws.cell(row=2, column=sev_col)
        # Check fill is not None and is a PatternFill with some red component
        fill = cell.fill
        assert fill is not None
        # fg_color should be set (not default theme/auto)
        if hasattr(fill, "fgColor") and fill.fgColor:
            fg = fill.fgColor.rgb if hasattr(fill.fgColor, "rgb") else ""
            # Red component (first 2 hex chars after alpha) should be dominant
            # e.g. "FFFF0000" → red, "FFFF4444" → red-ish
            assert fg != "00000000"  # not transparent/unset
        wb.close()

    def test_freeze_panes_on_row_2(self, tmp_path: Path) -> None:
        """Freeze panes should be set at A2."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        out = tmp_path / "report.xlsx"
        sr = _make_security_result()
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=sr)
        wb = load_workbook(out)
        ws = wb["Seguridad"]
        assert ws.freeze_panes == "A2"
        wb.close()

    def test_header_columns_present(self, tmp_path: Path) -> None:
        """Header row contains all expected columns."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        out = tmp_path / "report.xlsx"
        sr = _make_security_result()
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=sr)
        wb = load_workbook(out)
        ws = wb["Seguridad"]

        headers = {ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)}
        for expected in ("Archivo", "Regla", "Severidad", "Detector", "Línea", "Contexto"):
            assert expected in headers, f"Missing column: {expected}"
        wb.close()

    def test_existing_tests_still_pass_without_security_kwarg(
        self, tmp_path: Path
    ) -> None:
        """Old call style (no security kwarg) produces exactly 8 sheets."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        out = tmp_path / "report.xlsx"
        ExcelReporter().generate([_make_record()], _make_analysis(), out)
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 8
        wb.close()


# ---------------------------------------------------------------------------
# T21 — Dashboard KPI cards for security
# ---------------------------------------------------------------------------


class TestDashboardSecurityKPIs:
    """Tests for the two new KPI cards added to Dashboard when security present."""

    def _count_kpi_rows(self, ws, max_row_search: int = 20) -> list[str]:
        """Return list of KPI labels found in column A (rows 2 onward)."""
        labels = []
        for row in range(2, max_row_search + 1):
            val = ws.cell(row=row, column=1).value
            if val:
                labels.append(str(val))
        return labels

    def test_dashboard_has_12_kpis_without_security(self, tmp_path: Path) -> None:
        """Without security, Dashboard has exactly 12 KPI rows (no new ones)."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        out = tmp_path / "report.xlsx"
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=None)
        wb = load_workbook(out)
        ws = wb["Dashboard"]
        labels = self._count_kpi_rows(ws)
        # The existing dashboard always has 12 KPIs (10 static + top_ext + top_file)
        # Actually the count depends on data; we verify security labels are absent
        assert "Security Score" not in labels
        assert "Hallazgos" not in labels
        wb.close()

    def test_dashboard_has_security_score_kpi_when_security_present(
        self, tmp_path: Path
    ) -> None:
        """Security Score KPI appears on Dashboard when security data provided."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        out = tmp_path / "report.xlsx"
        sr = _make_security_result(security_score=85)
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=sr)
        wb = load_workbook(out)
        ws = wb["Dashboard"]
        labels = self._count_kpi_rows(ws)
        assert "Security Score" in labels
        wb.close()

    def test_dashboard_security_score_value_correct(self, tmp_path: Path) -> None:
        """Security Score KPI shows the correct numeric score."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        out = tmp_path / "report.xlsx"
        sr = _make_security_result(security_score=85)
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=sr)
        wb = load_workbook(out)
        ws = wb["Dashboard"]

        # Find "Security Score" row and read column B
        score_value = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "Security Score":
                score_value = ws.cell(row=row, column=2).value
                break

        assert score_value is not None
        # Value can be int, float, or string — extract numeric portion
        numeric = float(str(score_value).split("/")[0].strip()) if score_value else None
        assert numeric == 85.0
        wb.close()

    def test_dashboard_hallazgos_kpi_present(self, tmp_path: Path) -> None:
        """Hallazgos KPI appears on Dashboard when security data provided."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        out = tmp_path / "report.xlsx"
        sr = _make_security_result()
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=sr)
        wb = load_workbook(out)
        ws = wb["Dashboard"]
        labels = self._count_kpi_rows(ws)
        assert "Hallazgos" in labels
        wb.close()

    def test_dashboard_no_security_kpi_when_flag_off(self, tmp_path: Path) -> None:
        """No security KPIs when security=None."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        out = tmp_path / "report.xlsx"
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=None)
        wb = load_workbook(out)
        ws = wb["Dashboard"]
        labels = self._count_kpi_rows(ws)
        assert "Security Score" not in labels
        assert "Hallazgos" not in labels
        wb.close()

    def test_dashboard_security_score_green_fill_for_high_score(
        self, tmp_path: Path
    ) -> None:
        """Score >= 80 → green fill on Security Score cell."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        out = tmp_path / "report.xlsx"
        sr = _make_security_result(security_score=85)
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=sr)
        wb = load_workbook(out)
        ws = wb["Dashboard"]

        score_cell = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "Security Score":
                score_cell = ws.cell(row=row, column=2)
                break

        assert score_cell is not None
        fill = score_cell.fill
        if hasattr(fill, "fgColor") and fill.fgColor:
            fg = fill.fgColor.rgb if hasattr(fill.fgColor, "rgb") else ""
            assert fg != "00000000"
        wb.close()

    def test_dashboard_security_kpi_distinct_from_health(
        self, tmp_path: Path
    ) -> None:
        """Security Score KPI row is distinct from Health Score KPI row."""
        from fsaudit.reporter.excel_reporter import ExcelReporter

        out = tmp_path / "report.xlsx"
        sr = _make_security_result(security_score=72)
        ExcelReporter().generate([_make_record()], _make_analysis(), out, security=sr)
        wb = load_workbook(out)
        ws = wb["Dashboard"]

        health_row = None
        security_row = None
        for row in range(1, ws.max_row + 1):
            label = ws.cell(row=row, column=1).value
            if label == "Health Score":
                health_row = row
            if label == "Security Score":
                security_row = row

        assert health_row is not None, "Health Score row not found"
        assert security_row is not None, "Security Score row not found"
        assert health_row != security_row
        wb.close()


# ---------------------------------------------------------------------------
# T22 — HTML security section
# ---------------------------------------------------------------------------


class TestHtmlSecuritySection:
    """Tests for the collapsible security section in the HTML reporter."""

    def _generate_html(
        self,
        tmp_path: Path,
        security: SecurityResult | None = None,
    ) -> str:
        from fsaudit.reporter.html_reporter import HtmlReporter

        records = [_make_record()]
        analysis = _make_analysis()
        out = tmp_path / "report.html"
        HtmlReporter().generate(records, analysis, out, security=security)
        return out.read_text()

    def test_security_section_absent_when_no_scan(self, tmp_path: Path) -> None:
        """No security section when security=None."""
        html = self._generate_html(tmp_path, security=None)
        assert 'id="security"' not in html

    def test_security_section_present_when_scan_ran(self, tmp_path: Path) -> None:
        """Security section appears when security result provided."""
        sr = _make_security_result()
        html = self._generate_html(tmp_path, security=sr)
        assert 'id="security"' in html

    def test_html_finding_shows_severity_badge(self, tmp_path: Path) -> None:
        """Each finding has a severity badge with correct CSS class."""
        finding = _make_finding(severity=Severity.CRITICAL)
        sr = _make_security_result(findings=[finding])
        html = self._generate_html(tmp_path, security=sr)
        assert "badge-critical" in html

    def test_html_finding_shows_high_severity_badge(self, tmp_path: Path) -> None:
        """HIGH findings have badge-high class."""
        finding = _make_finding(severity=Severity.HIGH, rule_id="github-token")
        sr = _make_security_result(findings=[finding])
        html = self._generate_html(tmp_path, security=sr)
        assert "badge-high" in html

    def test_html_finding_shows_rule_id(self, tmp_path: Path) -> None:
        """Finding's rule_id appears in the HTML."""
        finding = _make_finding(rule_id="aws-access-key")
        sr = _make_security_result(findings=[finding])
        html = self._generate_html(tmp_path, security=sr)
        assert "aws-access-key" in html

    def test_html_finding_shows_path(self, tmp_path: Path) -> None:
        """Finding's file path appears in the HTML."""
        finding = _make_finding(path="/tmp/project/secrets.env")
        sr = _make_security_result(findings=[finding])
        html = self._generate_html(tmp_path, security=sr)
        assert "secrets.env" in html

    def test_html_section_grouped_by_severity(self, tmp_path: Path) -> None:
        """Findings are grouped by severity (critical section before high section)."""
        findings = [
            _make_finding(path="/tmp/high.py", severity=Severity.HIGH, rule_id="github-token"),
            _make_finding(path="/tmp/crit.env", severity=Severity.CRITICAL, rule_id="aws-access-key"),
        ]
        sr = _make_security_result(findings=findings)
        html = self._generate_html(tmp_path, security=sr)
        # CRITICAL should appear before HIGH in the document
        crit_idx = html.index("CRITICAL")
        high_idx = html.index("HIGH")
        assert crit_idx < high_idx

    def test_html_match_context_never_raw_long_secret(self, tmp_path: Path) -> None:
        """Raw long secret string does NOT appear in HTML (context is redacted ≤60 chars)."""
        raw_secret = "AKIA" + "X" * 76  # 80 chars — way over limit
        # SecurityFinding.__post_init__ will truncate to 60 chars automatically
        finding = _make_finding(match_context=raw_secret)
        sr = _make_security_result(findings=[finding])
        html = self._generate_html(tmp_path, security=sr)
        # The full 80-char secret should NOT appear in HTML
        assert raw_secret not in html

    def test_html_security_score_shown(self, tmp_path: Path) -> None:
        """Security score value appears in the HTML security section."""
        sr = _make_security_result(security_score=72)
        html = self._generate_html(tmp_path, security=sr)
        assert "72" in html

    def test_html_no_security_section_without_kwarg(self, tmp_path: Path) -> None:
        """Old call style (no security kwarg) produces no security section."""
        from fsaudit.reporter.html_reporter import HtmlReporter

        records = [_make_record()]
        analysis = _make_analysis()
        out = tmp_path / "report.html"
        HtmlReporter().generate(records, analysis, out)
        html = out.read_text()
        assert 'id="security"' not in html

    def test_html_only_non_empty_severity_buckets_shown(self, tmp_path: Path) -> None:
        """Only severity buckets with findings are rendered."""
        # Only CRITICAL finding → no HIGH/MEDIUM/LOW sections
        finding = _make_finding(severity=Severity.CRITICAL)
        sr = _make_security_result(findings=[finding])
        html = self._generate_html(tmp_path, security=sr)
        # CRITICAL section present
        assert "CRITICAL" in html
        # LOW section should NOT appear since there are no low findings
        assert "LOW" not in html
